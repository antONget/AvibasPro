from openpyxl import load_workbook
import logging


async def get_boarding_receipt(dict_check_ticket: dict, user_id: int) -> None:
    logging.info('get_boarding_receipt()')
    workbook = load_workbook(filename="services/roter_check.xlsx")
    sheet = workbook.active
    for key, value in dict_check_ticket.items():
        if key == 'H2':
            sheet[key].value = value
        else:

            sheet.merge_cells(key)
            top_left_cell = sheet[key.split(':')[0]]
            top_left_cell.value = value

    workbook.save(filename=f"TICKET/{user_id}.xlsx")
#
# data = {'B5:B6': 'Санкт-Петербург АВ № 2 — Вознесенье 895 , Пригородное',
#   'B11:B12': '00000000252010, тариф Пассажирский, заказ 00000000252, оплачен 27.12.2024 23:52',
#   'B14:B15': 'Иванов Сергей Игоревич, Паспорт гражданина РФ №12 34 123456',
#   'B17:B18': '400.0 руб.',
#   'E1:H1': '00000000252010',
#   'H2': 'Место 13',
#   'F5:F6': '12:20',
#   'G5:H6': '28.12.2024',
#   'F7:H9': 'Сясьстрой АС',
#   'F10:H13': 'ZHONG TONG LCK6127H, н225ук198\nЛенинградская область, Волховский район, Сясьстройское городское поселение, Сясьстрой, Петрозаводская ул, 14Б\nNone',
#   'F17:F18': '12:20',
#   'G17:H17': '28.12.2024',
#   'F19:H21': 'Лодейное Поле АС',
#   'F22:H24': 'Ленинградская область, Волховский район, Сясьстройское городское поселение, Сясьстрой, Петрозаводская ул, 14Б\nNone'
# }
#
# if __name__ == '__main__':
#     asyncio.run(get_boarding_receipt(dict_check_ticket=data))