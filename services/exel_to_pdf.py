from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib import colors

# Регистрируем шрифты, поддерживающие кириллицу
pdfmetrics.registerFont(TTFont("Arial", "arial(1).ttf"))  # Обычный Arial
pdfmetrics.registerFont(TTFont("Arial-Bold", "Arial.ttf"))  # Жирный Arial


def draw_dashed_vertical_line(c, x, y_start, y_end, dash_length=5, gap_length=3, line_width=1, color=colors.black):
    """
    Рисует вертикальную прерывистую линию на PDF.

    :param c: Canvas объекта PDF
    :param x: X-координата линии
    :param y_start: Начальная Y-координата линии
    :param y_end: Конечная Y-координата линии
    :param dash_length: Длина каждого штриха
    :param gap_length: Длина промежутка между штрихами
    :param line_width: Толщина линии
    :param color: Цвет линии
    """
    # Настройка цвета и толщины линии
    c.setStrokeColor(color)
    c.setLineWidth(line_width)

    # Рисование пунктирной линии
    y = y_start
    while y > y_end:
        c.line(x, y, x, max(y - dash_length, y_end))
        y -= dash_length + gap_length


def wrap_text(c, text, x, y, max_width, font_name, font_size, line_height):
    """
    Разбивает текст на строки, чтобы он помещался в заданной ширине.

    :param c: Canvas для рисования
    :param text: Текст для вывода
    :param x: Начальная координата X
    :param y: Начальная координата Y
    :param max_width: Максимальная ширина строки
    :param font_name: Имя шрифта
    :param font_size: Размер шрифта
    :param line_height: Высота строки
    """
    words = text.split(' ')
    line = ""
    for word in words:
        test_line = f"{line} {word}".strip()
        if c.stringWidth(test_line, font_name, font_size) <= max_width:
            line = test_line
        else:
            c.drawString(x, y, line)
            y -= line_height
            line = word
    if line:
        c.drawString(x, y, line)
    return y


def excel_to_pdf(input_file, output_file):
    """
    Конвертирует Excel в PDF с разным позиционированием текста и корректным учетом высоты ячейки.

    :param input_file: Путь к файлу Excel
    :param output_file: Путь к файлу PDF для сохранения результата
    """
    # Загрузка Excel файла
    wb = load_workbook(input_file)
    sheet = wb.active  # Выбираем активный лист

    # Создаем PDF
    c = canvas.Canvas(output_file, pagesize=letter)
    width, height = letter  # Размер страницы

    # Начальные координаты
    left_y_position = height - 40  # Начальная высота для левой части
    right_y_position = height - 40  # Начальная высота для правой части
    line_height = 12  # Стандартная высота строки для переноса текста

    # Координаты для левого и правого текста
    left_start_x = 40  # Левый текст
    right_start_x = width - 300  # Правый текст (задается фиксированным отступом от правого края)
    y = []
    for row in sheet.iter_rows():
        left_x_position = left_start_x  # Текущая координата X для левого текста
        right_x_position = right_start_x  # Текущая координата X для правого текста

        # Флаги наличия текста в каждой части
        left_has_text = False
        right_has_text = False

        for cell in row:
            # Определяем текущий столбец
            column_letter = get_column_letter(cell.column)
            # Получаем текст из ячейки
            cell_text = str(cell.value) if cell.value else ""


            # Определяем ширину столбца

            if column_letter == "A" and cell_text and not row[cell.column].value:
                next_column_width = sheet.column_dimensions["B"].width if "B" in sheet.column_dimensions else 10
                column_width = (sheet.column_dimensions["A"].width + next_column_width) * 5
            elif column_letter == "G" and cell_text and not row[cell.column].value:
                next_column_width = sheet.column_dimensions["H"].width if "H" in sheet.column_dimensions else 10
                column_width = (sheet.column_dimensions["G"].width + next_column_width) * 7
            elif column_letter == "E" and cell_text and not row[cell.column].value:
                next_column_width = sheet.column_dimensions["D"].width if "D" in sheet.column_dimensions else 10
                column_width = (sheet.column_dimensions["E"].width + next_column_width) * 7
            # Проверка для объединения ячеек G и H
            elif column_letter == "F" and not row[cell.column].value:
                next_column_width = sheet.column_dimensions["H"].width if "H" in sheet.column_dimensions else 10
                next_column_width_1 = sheet.column_dimensions["G"].width if "G" in sheet.column_dimensions else 10
                column_width = (sheet.column_dimensions["F"].width + next_column_width + next_column_width_1) * 4

            else:
                # Обычная ширина столбца
                column_width = sheet.column_dimensions[
                                   column_letter].width * 5 if column_letter in sheet.column_dimensions else 70
            if left_x_position + column_width > 300 and column_letter in ['A', 'B', 'C']:
                column_width = 300 - left_x_position
            # Устанавливаем шрифт
            if cell.font.bold:
                font_name = "Arial-Bold"
            else:
                font_name = "Arial"
            font_size = cell.font.size or 10
            c.setFont(font_name, font_size)

            # Устанавливаем цвет текста
            font_color = cell.font.color
            if font_color and font_color.rgb:
                hex_color = font_color.rgb[-6:]  # Последние 6 символов (RRGGBB)
                c.setFillColor(colors.HexColor(f"#{hex_color}"))
            else:
                c.setFillColor(colors.black)

            # Печатаем текст с переносом
            if cell_text:

                if column_letter in ['A', 'B', 'C']:  # Столбцы слева
                    left_has_text = True  # Текст есть в левой части
                    new_y_position = wrap_text(
                        c, cell_text, left_x_position, left_y_position, column_width, font_name, font_size, line_height
                    )
                    left_y_position = min(left_y_position, new_y_position)  # Обновляем Y координату
                    left_x_position += column_width
                elif column_letter in ['D', 'E', 'F', 'G', 'H']:  # Столбцы справа
                    if "●" in cell_text:
                        y.append(right_y_position)
                    if "▼" in cell_text:
                        right_x_position -=2
                        y.append(right_y_position)
                    if "Место" in cell_text:
                        text_width = c.stringWidth(cell_text, font_name, font_size)
                        # Проверяем, помещается ли текст в пределах max_x
                        max_x = 560
                        if right_x_position + text_width > max_x:
                            # Если не помещается, сдвигаем текст влево
                            right_x_position = max_x - text_width
                        right_has_text = True  # Текст есть в правой части
                        new_y_position = wrap_text(
                            c, cell_text, right_x_position, right_y_position, column_width, font_name, font_size, line_height
                        )
                        right_y_position = min(right_y_position, new_y_position)  # Обновляем Y координату
                        right_x_position += column_width

                    elif ' ' in cell_text:
                        right_has_text = True
                        lines = cell_text.split("\n")
                        for line in lines:
                            new_y_position = wrap_text(
                                c, line, right_x_position, right_y_position, column_width, font_name, font_size,
                                line_height
                            )
                            right_y_position = min(right_y_position, new_y_position)  # Обновляем Y координату

                            right_y_position -= line_height
                    else:
                        right_has_text = True  # Текст есть в правой части
                        new_y_position = wrap_text(
                            c, cell_text, right_x_position, right_y_position, column_width, font_name, font_size,
                            line_height
                        )
                        right_y_position = min(right_y_position, new_y_position)  # Обновляем Y координату
                        right_x_position += column_width
            else:
                if column_letter in ['A', 'B', 'C']:
                    left_x_position += column_width
                elif column_letter in ['D', 'E', 'F', 'G', 'H']:
                    right_x_position += column_width
                    # Определяем минимальное Y для перехода на следующую строку
        # Если есть текст в соответствующей части, используем рассчитанный Y, иначе стандартное смещение
        if  left_has_text:
            left_y_position -= 24
        if  right_has_text:
            right_y_position -= 24

        # Берем минимальную координату между левой и правой частью
        min_y_position = min(left_y_position, right_y_position)

        # Если позиция текста выходит за пределы страницы, добавляем новую страницу
        if left_y_position < 40 or right_y_position < 40:
            c.showPage()
            left_y_position = right_y_position = height - 40  # Возвращаемся на верх страницы

    x = 315  # X-координата
    y_start = y[0] - 10  # Верхняя точка линии
    y_end = y[1] - 10  # Нижняя точка линии

    draw_dashed_vertical_line(c, x = x, y_start = y_start, y_end = y_end)
    # Сохраняем PDF
    c.save()


# Пример использования
# excel_to_pdf("843554518.xlsx", "output.pdf")
